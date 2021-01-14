# -*- coding: utf-8 -*-

# импорт необходимых библиотек
from pandas import read_excel
from collections import Counter, defaultdict
from openpyxl.reader import excel


# открываем на чтение файл
logs_list = read_excel('logs.xlsx', sheet_name='log')
# считываем логи
logs = logs_list.to_dict(orient='records')
# открываем файл на запись
wb = excel.load_workbook('report.xlsx')
sheet = wb.active

# этот словарь предназначен для хранения кол-ва посещений. Формат ключа: <Имя браузера> + '-' + <Номер месяца>
visits = defaultdict(int)
# этот массив предназначен для подсчета общего кол-ва посещений для каждого браузера
raw_browsers_data = []
# этот массив будет хранить названия самых популярных браузеров
most_common_browsers = []

# здесь итерируемся по логам и заполняем значениями два ранее созданных массива
for record in logs:
    browser = record['Браузер']
    raw_browsers_data.append(browser)
    visits[browser + '-' + str(record['Дата посещения'].month)] += 1

# здесь используем заполненый ранее массив для подсчета самых популярных браузеров и кол-ва посещений для них
# далее очищаем массив для экономии оперативной памяти
browsers_visits_count = Counter(raw_browsers_data)
raw_browsers_data.clear()

# здесь итерируемся по массиву самых популярных браузеров и заполняем ячейки в таблице
# также заполняем массив самых популярных браузеров
for i in range(5, 12):
    most_common_value = browsers_visits_count.most_common(7)[i - 5]
    sheet.cell(row=i, column=1).value = most_common_value[0]
    sheet.cell(row=i, column=2).value = most_common_value[1]
    most_common_browsers.append(most_common_value)

# здесь итерируемся по месяцам и по массиву самых популярных браузеров и заполняем ячейки посещений
# по месяцам для 7ми самых популярных браузеров
for i in range(1, 13):
    sumPerMonth = 0
    for a in range(7):
        visitsPerMonth = visits[str(most_common_browsers[a][0]) + '-' + str(i)]
        sheet.cell(row=(a + 5), column=(i + 2)).value = visitsPerMonth
        sumPerMonth += visitsPerMonth
    sheet.cell(row=12, column=(i + 2)).value = sumPerMonth

# здесь заканчивается секция связанная с браузерами

# создаем массив для хранения всех товаров
most_popular_items = []

# заполняем ранее созданный массив
for record in logs:
    most_popular_items += record['Купленные товары'].split(sep=',')

# выделяем самые популярные товары и удаляем елементы 'Еще 2 варианта' и 'Еще 3 варианта'
most_popular_items_count = Counter(most_popular_items).most_common(8)
for i in most_popular_items_count:
    if i[0] == 'Ещё 2 варианта' or i[0] == 'Ещё 3 варианта':
        del most_popular_items_count[most_popular_items_count.index(i)]

# заполняем ячейки с самыми популярными товарами
for i in range(1, 8):
    sheet.cell(row=(i + 18), column=1).value = most_popular_items_count[i - 1][0]
    sheet.cell(row=(i + 18), column=2).value = most_popular_items_count[i - 1][1]

# создаем словарь наподобии того, что использовался ранее для подсчета кол-ва посещений
# за месяц для браузеров
popular_items_month = defaultdict(int)

# тут заполняем словарь
for most_pop_it in most_popular_items_count:
    for record in logs:
        cur_it = record['Купленные товары'].split(',')
        if most_pop_it[0] in cur_it:
            popular_items_month[most_pop_it[0] + '-' + str(record['Дата посещения'].month)] += 1

# создаем словарь продаж по месяцам (общим)
most_pop_it_sales_month = defaultdict(int)

# тут заполняем ячейки значениями и одновременно заполняем словарь для общих продаж за месяц
for a in range(7):
    for b in range(12):
        value_per_month = popular_items_month[most_popular_items_count[a][0] + '-' + str(b + 1)]
        sheet.cell(row=(a + 19), column=(b + 3)).value = value_per_month
        most_pop_it_sales_month[b] += value_per_month

# заполняем ячейки таблицы из заполненного ранее словаря
for i in range(12):
    sheet.cell(row=26, column=(i + 3)).value = most_pop_it_sales_month[i]

# теперь создаем два массива для предпочтений по полам
most_pop_it_male = []
most_pop_it_female = []

# заполняем эти два массива одновременно фильтруя инвалидные элементы
for record in logs:
    pur_it = record['Купленные товары'].split(',')
    if not ('Ещё 2 варианта' in pur_it) and not ('Ещё 3 варианта' in pur_it):
        if record['Пол'] == 'м':
            most_pop_it_male += pur_it
        else:
            most_pop_it_female += pur_it

# тут по классической схеме
# создаем счетчики
most_pop_it_male_cont = Counter(most_pop_it_male)
most_pop_it_female_cont = Counter(most_pop_it_female)

# высчитываем их длину
most_pop_it_male_cont_len = len(most_pop_it_male_cont)
most_pop_it_female_cont_len = len(most_pop_it_female_cont)

# заполняем ячейки самыми популярными товарами по полам
sheet.cell(row=31, column=2).value = most_pop_it_male_cont.most_common(1)[0][0]
sheet.cell(row=32, column=2).value = most_pop_it_female_cont.most_common(1)[0][0]

# и самыми невостребованными
sheet.cell(row=33, column=2).value = most_pop_it_male_cont.most_common()[:-(most_pop_it_male_cont_len + 1):-1][0][0]
sheet.cell(row=34, column=2).value = most_pop_it_female_cont.most_common()[:-(most_pop_it_female_cont_len + 1):-1][0][0] #показываем 2ой элемент т. к. первый нужно удалить

# сохраняем итог!
wb.save('report.xlsx')
